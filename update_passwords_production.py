"""
Скрипт для обновления паролей студентов из Excel файла на production сервере.

Как использовать:
1. Загрузить Excel файл на сервер:
   scp "База ученики.xlsx" jsi:~/students.xlsx

2. Загрузить этот скрипт на сервер:
   scp update_passwords_production.py jsi:~/update_passwords.py

3. Скопировать файлы в контейнер:
   ssh jsi "cd ~/english-crm && sudo docker compose cp ~/students.xlsx backend:/app/students.xlsx"
   ssh jsi "cd ~/english-crm && sudo docker compose cp ~/update_passwords.py backend:/app/update_passwords.py"

4. Запустить скрипт:
   ssh jsi "cd ~/english-crm && sudo docker compose exec backend python /app/update_passwords.py"
"""
import asyncio
import bcrypt
from openpyxl import load_workbook
from sqlalchemy import select
from app.models.user import User
from app.database import async_session_maker


def hash_password(password: str) -> str:
    """Хеширует пароль с помощью bcrypt"""
    pwd_bytes = password.encode('utf-8')
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(pwd_bytes, salt)
    return hashed.decode('utf-8')


async def update_passwords():
    """Обновляет пароли студентов из Excel файла"""

    excel_path = "/app/students.xlsx"

    # Читаем Excel файл
    print(f"Reading file: {excel_path}")
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    # Получаем заголовки
    headers = [cell.value for cell in ws[1]]
    print(f"Columns: {headers}")

    # Email в колонке B (индекс 1), Пароль в колонке C (индекс 2)
    email_col_idx = 1  # Колонка B (0-based: A=0, B=1, C=2)
    password_col_idx = 2  # Колонка C

    print(f"Email column: {headers[email_col_idx]}")
    print(f"Password column: {headers[password_col_idx]}")

    # Собираем данные из Excel
    students_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
        if len(row) > password_col_idx:
            email = row[email_col_idx]
            password = row[password_col_idx]
            if email and password:
                students_data.append((str(email).strip().lower(), str(password).strip()))

    wb.close()

    print(f"Total students to update: {len(students_data)}")

    success_count = 0
    not_found_count = 0
    error_count = 0
    not_found_emails = []

    print("\nStarting password update...\n")

    async with async_session_maker() as session:
        for email, password in students_data:

            try:
                # Находим студента по email
                result = await session.execute(
                    select(User).where(User.email == email, User.role == 'student')
                )
                user = result.scalar_one_or_none()

                if not user:
                    not_found_emails.append(email)
                    not_found_count += 1
                    continue

                # Хешируем новый пароль
                hashed_password = hash_password(password)

                # Обновляем пароль
                user.password_hash = hashed_password
                await session.commit()

                print(f"[OK] Updated: {user.name} ({email})")
                success_count += 1

            except Exception as e:
                print(f"[ERROR] For {email}: {str(e)}")
                error_count += 1
                await session.rollback()

    # Итоговая статистика
    print("\n" + "="*60)
    print("UPDATE SUMMARY:")
    print(f"Successfully updated: {success_count}")
    print(f"Not found in DB: {not_found_count}")
    print(f"Errors: {error_count}")
    print("="*60)

    if not_found_emails:
        print(f"\nStudents not found in DB ({len(not_found_emails)} total):")
        for email in not_found_emails[:30]:  # Показываем первые 30
            print(f"  - {email}")
        if len(not_found_emails) > 30:
            print(f"  ... and {len(not_found_emails) - 30} more")


if __name__ == "__main__":
    asyncio.run(update_passwords())
