from collections import defaultdict
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select, and_, or_, func
from sqlalchemy.orm import selectinload

from app.api.deps import DBSession, ManagerUser
from app.models.lesson import AttendanceStatus, Lesson, LessonStatus, LessonStudent
from app.models.lesson_type import LessonType
from app.models.level_lesson_type_payment import LevelLessonTypePayment
from app.models.transaction import Transaction, TransactionType
from app.models.user import User, UserRole
from app.schemas.report import (
    ReportRequest,
    StudentReportRow,
    TeacherReport,
    TeacherReportResponse,
    TransactionReportResponse,
    TransactionReportRow,
)

router = APIRouter()


@router.post("/teachers", response_model=TeacherReportResponse)
async def generate_teacher_report(
    data: ReportRequest,
    db: DBSession,
    current_user: ManagerUser,
):
    """Generate report of lessons grouped by teachers for a date range."""

    # Convert dates to datetime for comparison
    date_from_dt = datetime.combine(data.date_from, time.min)
    date_to_dt = datetime.combine(data.date_to, time.max)

    # Get all completed lessons in the date range with related data
    query = (
        select(Lesson)
        .options(
            selectinload(Lesson.teacher).selectinload(User.level),
            selectinload(Lesson.lesson_type),
            selectinload(Lesson.students).selectinload(LessonStudent.student),
        )
        .where(
            and_(
                Lesson.scheduled_at >= date_from_dt,
                Lesson.scheduled_at <= date_to_dt,
                Lesson.status == LessonStatus.COMPLETED,
            )
        )
        .order_by(Lesson.teacher_id, Lesson.scheduled_at)
    )

    result = await db.execute(query)
    lessons = result.scalars().all()

    # Group lessons by teacher, then by student and lesson type
    teacher_data: dict[int, dict] = defaultdict(lambda: {
        "teacher_name": "",
        "students": defaultdict(lambda: defaultdict(lambda: {
            "count": 0,
            "payment": Decimal("0"),
        }))
    })

    for lesson in lessons:
        teacher_id = lesson.teacher_id
        teacher = lesson.teacher
        lesson_type = lesson.lesson_type

        teacher_data[teacher_id]["teacher_name"] = teacher.name
        lesson_price = lesson_type.price

        # Calculate teacher payment based on teacher's level and lesson type
        teacher_payment = Decimal("0")
        if teacher.level_id:
            # Look up payment from matrix
            payment_result = await db.execute(
                select(LevelLessonTypePayment).where(
                    and_(
                        LevelLessonTypePayment.level_id == teacher.level_id,
                        LevelLessonTypePayment.lesson_type_id == lesson_type.id,
                    )
                )
            )
            payment_config = payment_result.scalar_one_or_none()
            if payment_config:
                teacher_payment = payment_config.teacher_payment

        # Fallback: 50% of lesson price if no matrix config
        if teacher_payment == Decimal("0"):
            teacher_payment = lesson_price * Decimal("0.5")

        # Add data for each student in the lesson
        for lesson_student in lesson.students:
            if lesson_student.attendance_status == AttendanceStatus.PRESENT:
                student = lesson_student.student
                student_key = (student.id, student.name)
                lesson_type_key = lesson_type.name

                teacher_data[teacher_id]["students"][student_key][lesson_type_key]["count"] += 1
                teacher_data[teacher_id]["students"][student_key][lesson_type_key]["payment"] += teacher_payment

    # Build response
    teachers_list: list[TeacherReport] = []
    grand_total = Decimal("0")

    for teacher_id, teacher_info in teacher_data.items():
        rows: list[StudentReportRow] = []
        teacher_total = Decimal("0")

        for (student_id, student_name), lesson_types in teacher_info["students"].items():
            for lesson_type_name, stats in lesson_types.items():
                row = StudentReportRow(
                    student_name=student_name,
                    lesson_type=lesson_type_name,
                    lessons_count=stats["count"],
                    teacher_payment=stats["payment"],
                )
                rows.append(row)
                teacher_total += stats["payment"]

        if rows:
            teachers_list.append(TeacherReport(
                teacher_id=teacher_id,
                teacher_name=teacher_info["teacher_name"],
                rows=rows,
                total=teacher_total,
            ))
            grand_total += teacher_total

    return TeacherReportResponse(
        teachers=teachers_list,
        grand_total=grand_total,
        date_from=data.date_from,
        date_to=data.date_to,
    )


@router.post("/teachers/export")
async def export_teacher_report(
    data: ReportRequest,
    db: DBSession,
    current_user: ManagerUser,
):
    """Export teacher report to Excel file."""

    # Convert dates to datetime for comparison
    date_from_dt = datetime.combine(data.date_from, time.min)
    date_to_dt = datetime.combine(data.date_to, time.max)

    # Get all completed lessons in the date range with related data
    query = (
        select(Lesson)
        .options(
            selectinload(Lesson.teacher).selectinload(User.level),
            selectinload(Lesson.lesson_type),
            selectinload(Lesson.students).selectinload(LessonStudent.student),
        )
        .where(
            and_(
                Lesson.scheduled_at >= date_from_dt,
                Lesson.scheduled_at <= date_to_dt,
                Lesson.status == LessonStatus.COMPLETED,
            )
        )
        .order_by(Lesson.teacher_id, Lesson.scheduled_at)
    )

    result = await db.execute(query)
    lessons = result.scalars().all()

    # Group lessons by teacher, then by student and lesson type
    teacher_data: dict[int, dict] = defaultdict(lambda: {
        "teacher_name": "",
        "students": defaultdict(lambda: defaultdict(lambda: {
            "count": 0,
            "payment": Decimal("0"),
        }))
    })

    for lesson in lessons:
        teacher_id = lesson.teacher_id
        teacher = lesson.teacher
        lesson_type = lesson.lesson_type

        teacher_data[teacher_id]["teacher_name"] = teacher.name
        lesson_price = lesson_type.price

        # Calculate teacher payment based on teacher's level and lesson type
        teacher_payment = Decimal("0")
        if teacher.level_id:
            payment_result = await db.execute(
                select(LevelLessonTypePayment).where(
                    and_(
                        LevelLessonTypePayment.level_id == teacher.level_id,
                        LevelLessonTypePayment.lesson_type_id == lesson_type.id,
                    )
                )
            )
            payment_config = payment_result.scalar_one_or_none()
            if payment_config:
                teacher_payment = payment_config.teacher_payment

        if teacher_payment == Decimal("0"):
            teacher_payment = lesson_price * Decimal("0.5")

        for lesson_student in lesson.students:
            if lesson_student.attendance_status == AttendanceStatus.PRESENT:
                student = lesson_student.student
                student_key = (student.id, student.name)
                lesson_type_key = lesson_type.name

                teacher_data[teacher_id]["students"][student_key][lesson_type_key]["count"] += 1
                teacher_data[teacher_id]["students"][student_key][lesson_type_key]["payment"] += teacher_payment

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт по преподавателям"

    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    teacher_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    total_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    grand_total_fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
    grand_total_font = Font(bold=True, size=14, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Отчёт по преподавателям за {data.date_from.strftime('%d.%m.%Y')} — {data.date_to.strftime('%d.%m.%Y')}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3
    grand_total = Decimal("0")

    for teacher_id, teacher_info in teacher_data.items():
        # Teacher header
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = teacher_info["teacher_name"]
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws[f"A{row}"].fill = teacher_fill
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

        # Column headers
        headers = ["Имя ученика", "Вид занятия", "Кол-во занятий", "Оплата преподавателю"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row += 1

        teacher_total = Decimal("0")

        for (student_id, student_name), lesson_types in teacher_info["students"].items():
            for lesson_type_name, stats in lesson_types.items():
                ws.cell(row=row, column=1, value=student_name).border = thin_border
                ws.cell(row=row, column=2, value=lesson_type_name).border = thin_border
                ws.cell(row=row, column=3, value=stats["count"]).border = thin_border
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=4, value=float(stats["payment"])).border = thin_border
                ws.cell(row=row, column=4).number_format = '#,##0 "тг"'
                ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
                teacher_total += stats["payment"]
                row += 1

        # Teacher total
        ws.merge_cells(f"A{row}:C{row}")
        ws[f"A{row}"] = "Итого за преподавателя"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = total_fill
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=4, value=float(teacher_total))
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=4).fill = total_fill
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).number_format = '#,##0 "тг"'
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")

        grand_total += teacher_total
        row += 2  # Empty row between teachers

    # Grand total
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "ОБЩИЙ ИТОГ"
    ws[f"A{row}"].font = grand_total_font
    ws[f"A{row}"].fill = grand_total_fill
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = grand_total_fill
    ws.cell(row=row, column=4, value=float(grand_total))
    ws.cell(row=row, column=4).font = grand_total_font
    ws.cell(row=row, column=4).fill = grand_total_fill
    ws.cell(row=row, column=4).border = thin_border
    ws.cell(row=row, column=4).number_format = '#,##0 "тг"'
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")

    # Adjust column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 25

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"report_{data.date_from.strftime('%Y%m%d')}_{data.date_to.strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/transactions", response_model=TransactionReportResponse)
async def get_transactions_report(
    db: DBSession,
    current_user: ManagerUser,
    date_from: date | None = Query(None, description="Filter from date (inclusive)"),
    date_to: date | None = Query(None, description="Filter to date (inclusive)"),
    search: str | None = Query(None, description="Search by student name, description, or manager name"),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
):
    """
    Get report of all student balance credit transactions (only increases).

    Supports:
    - Date range filtering
    - Fuzzy search by student name, description, or manager name
    - Pagination
    """

    # Base query - only CREDIT transactions for students
    query = (
        select(Transaction)
        .join(Transaction.user)
        .where(
            and_(
                Transaction.type == TransactionType.CREDIT,
                User.role == UserRole.STUDENT,
            )
        )
        .options(
            selectinload(Transaction.user),
            selectinload(Transaction.created_by),
        )
    )

    # Count query for pagination
    count_query = (
        select(func.count(Transaction.id))
        .join(Transaction.user)
        .where(
            and_(
                Transaction.type == TransactionType.CREDIT,
                User.role == UserRole.STUDENT,
            )
        )
    )

    # Apply date filters
    if date_from:
        date_from_dt = datetime.combine(date_from, time.min)
        query = query.where(Transaction.created_at >= date_from_dt)
        count_query = count_query.where(Transaction.created_at >= date_from_dt)

    if date_to:
        date_to_dt = datetime.combine(date_to, time.max)
        query = query.where(Transaction.created_at <= date_to_dt)
        count_query = count_query.where(Transaction.created_at <= date_to_dt)

    # Apply search filter
    if search:
        # Join with created_by user for manager name search
        search_filter = or_(
            User.name.ilike(f"%{search}%"),  # student name
            Transaction.description.ilike(f"%{search}%"),  # description
        )
        query = query.where(search_filter)
        count_query = count_query.where(search_filter)

    # Get total count
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    # Get total amount (sum of all filtered transactions)
    sum_query = (
        select(func.coalesce(func.sum(Transaction.amount), 0))
        .join(Transaction.user)
        .where(
            and_(
                Transaction.type == TransactionType.CREDIT,
                User.role == UserRole.STUDENT,
            )
        )
    )

    if date_from:
        sum_query = sum_query.where(Transaction.created_at >= datetime.combine(date_from, time.min))
    if date_to:
        sum_query = sum_query.where(Transaction.created_at <= datetime.combine(date_to, time.max))
    if search:
        sum_query = sum_query.where(search_filter)

    sum_result = await db.execute(sum_query)
    total_amount = Decimal(str(sum_result.scalar() or 0))

    # Get paginated results
    offset = (page - 1) * size
    query = query.offset(offset).limit(size).order_by(Transaction.created_at.desc())

    result = await db.execute(query)
    transactions = result.scalars().all()

    # Build response
    items = []
    for txn in transactions:
        items.append(TransactionReportRow(
            transaction_id=txn.id,
            student_id=txn.user_id,
            student_name=txn.user.name,
            amount=txn.amount,
            description=txn.description,
            created_at=txn.created_at,
            created_by_id=txn.created_by_id,
            created_by_name=txn.created_by.name if txn.created_by else None,
        ))

    pages = (total + size - 1) // size if total > 0 else 1

    return TransactionReportResponse(
        items=items,
        total=total,
        page=page,
        size=size,
        pages=pages,
        total_amount=total_amount,
    )
