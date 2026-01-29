"""
Import script for students, groups, and individual teacher-student relationships.

Usage:
    python -m scripts.import_data
"""

import asyncio
import csv
import sys
from decimal import Decimal
from pathlib import Path

import bcrypt
import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

# Add parent to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.database import async_session_maker
from app.models.user import User, UserRole
from app.models.group import Group, GroupStudent
from app.models.teacher_student import TeacherStudent


def hash_password(password: str) -> str:
    """Hash a password using bcrypt."""
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()


def generate_password(email: str) -> str:
    """Generate a simple password from email."""
    # Use first part of email + "123"
    local_part = email.split("@")[0]
    return f"{local_part}123"


async def get_or_create_user(
    session: AsyncSession,
    email: str,
    name: str = None,
    phone: str = None,
    balance: Decimal = Decimal("0.00"),
    role: UserRole = UserRole.STUDENT,
    password: str = None,
) -> User:
    """Get existing user or create new one."""
    email = email.strip().lower()

    # Check if user exists
    result = await session.execute(select(User).where(User.email == email))
    user = result.scalar_one_or_none()

    if user:
        return user

    # Create new user
    if not name:
        name = email.split("@")[0].replace(".", " ").replace("_", " ").title()

    if not password:
        password = generate_password(email)

    user = User(
        name=name.strip(),
        email=email,
        phone=phone,
        balance=balance,
        role=role,
        password_hash=hash_password(password),
        is_active=True,
    )
    session.add(user)
    await session.flush()
    return user


async def import_students(session: AsyncSession, file_path: str) -> dict:
    """Import students from Excel file."""
    print(f"\n=== Importing students from {file_path} ===")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    students_by_email = {}
    created = 0
    updated = 0
    errors = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        name, email, login, phone, balance, password = row[:6]

        if not email:
            continue

        email = str(email).strip().lower()

        try:
            # Clean phone number
            phone_str = None
            if phone:
                phone_str = str(int(phone)) if isinstance(phone, float) else str(phone)
                if not phone_str.startswith("+"):
                    phone_str = "+" + phone_str

            # Clean balance
            balance_dec = Decimal("0.00")
            if balance:
                try:
                    balance_dec = Decimal(str(balance))
                except:
                    balance_dec = Decimal("0.00")

            # Clean name
            name_str = str(name).strip() if name else email.split("@")[0]

            # Password - use from file or generate
            pwd = str(password).strip() if password else generate_password(email)

            user = await get_or_create_user(
                session,
                email=email,
                name=name_str,
                phone=phone_str,
                balance=balance_dec,
                role=UserRole.STUDENT,
                password=pwd,
            )

            students_by_email[email] = user

            if user.id:
                created += 1
            else:
                updated += 1

        except Exception as e:
            errors.append(f"Row {i}: {email} - {e}")

    wb.close()

    print(f"  Created/found: {created} students")
    if errors:
        print(f"  Errors: {len(errors)}")
        for err in errors[:5]:
            print(f"    {err}")

    return students_by_email


async def import_teachers_from_groups(session: AsyncSession, csv_path: str) -> dict:
    """Extract and create teachers from groups CSV."""
    print(f"\n=== Extracting teachers from {csv_path} ===")

    teachers_by_email = {}

    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f, delimiter=';')
        next(reader)  # Skip header

        for row in reader:
            if len(row) < 2:
                continue

            teacher_email = row[1].strip().lower() if row[1] else None

            if teacher_email and teacher_email not in teachers_by_email and '@' in teacher_email:
                user = await get_or_create_user(
                    session,
                    email=teacher_email,
                    role=UserRole.TEACHER,
                )
                teachers_by_email[teacher_email] = user

    print(f"  Found/created: {len(teachers_by_email)} teachers from groups")
    return teachers_by_email


async def import_teachers_from_individuals(session: AsyncSession, file_path: str, existing_teachers: dict) -> dict:
    """Extract and create teachers from individuals Excel."""
    print(f"\n=== Extracting teachers from {file_path} ===")

    teachers_by_email = existing_teachers.copy()

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        teacher_email = row[0]

        if teacher_email:
            teacher_email = str(teacher_email).strip().lower()

            if teacher_email not in teachers_by_email and '@' in teacher_email:
                user = await get_or_create_user(
                    session,
                    email=teacher_email,
                    role=UserRole.TEACHER,
                )
                teachers_by_email[teacher_email] = user

    wb.close()

    print(f"  Total teachers: {len(teachers_by_email)}")
    return teachers_by_email


async def import_groups(
    session: AsyncSession,
    csv_path: str,
    teachers: dict,
    students: dict,
) -> list:
    """Import groups from CSV file."""
    print(f"\n=== Importing groups from {csv_path} ===")

    groups_created = []
    errors = []

    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f, delimiter=';')
        next(reader)  # Skip header

        for i, row in enumerate(reader, 2):
            if len(row) < 3:
                continue

            group_name = row[0].strip() if row[0] else None
            teacher_email = row[1].strip().lower() if row[1] else None
            student_emails_str = row[2] if row[2] else ""

            if not group_name or group_name == "Название класса":
                continue

            try:
                # Get or create teacher
                teacher = None
                if teacher_email and teacher_email in teachers:
                    teacher = teachers[teacher_email]
                elif teacher_email and '@' in teacher_email:
                    teacher = await get_or_create_user(
                        session,
                        email=teacher_email,
                        role=UserRole.TEACHER,
                    )
                    teachers[teacher_email] = teacher

                # Check if group exists
                result = await session.execute(
                    select(Group).where(Group.name == group_name)
                )
                group = result.scalar_one_or_none()

                if not group:
                    group = Group(
                        name=group_name,
                        teacher_id=teacher.id if teacher else None,
                        is_active=True,
                    )
                    session.add(group)
                    await session.flush()

                # Parse student emails
                student_emails = [
                    e.strip().lower()
                    for e in student_emails_str.split(",")
                    if e.strip() and '@' in e.strip()
                ]

                # Add students to group
                for student_email in student_emails:
                    # Get or create student
                    if student_email in students:
                        student = students[student_email]
                    else:
                        student = await get_or_create_user(
                            session,
                            email=student_email,
                            role=UserRole.STUDENT,
                        )
                        students[student_email] = student

                    # Check if already in group
                    result = await session.execute(
                        select(GroupStudent).where(
                            GroupStudent.group_id == group.id,
                            GroupStudent.student_id == student.id,
                        )
                    )
                    existing = result.scalar_one_or_none()

                    if not existing:
                        gs = GroupStudent(
                            group_id=group.id,
                            student_id=student.id,
                        )
                        session.add(gs)

                groups_created.append(group)

            except Exception as e:
                errors.append(f"Row {i}: {group_name} - {e}")

    print(f"  Created/updated: {len(groups_created)} groups")
    if errors:
        print(f"  Errors: {len(errors)}")
        for err in errors[:5]:
            print(f"    {err}")

    return groups_created


async def create_individual_assignments(
    session: AsyncSession,
    file_path: str,
    teachers: dict,
    students: dict,
) -> int:
    """Create direct teacher-student assignments from Individuals file."""
    print(f"\n=== Creating individual teacher-student assignments from {file_path} ===")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    created = 0
    errors = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        teacher_email, student_email = row[:2]

        if not teacher_email or not student_email:
            continue

        teacher_email = str(teacher_email).strip().lower()
        student_email = str(student_email).strip().lower()

        if '@' not in teacher_email or '@' not in student_email:
            continue

        try:
            # Get or create teacher
            if teacher_email in teachers:
                teacher = teachers[teacher_email]
            else:
                teacher = await get_or_create_user(
                    session,
                    email=teacher_email,
                    role=UserRole.TEACHER,
                )
                teachers[teacher_email] = teacher

            # Get or create student
            if student_email in students:
                student = students[student_email]
            else:
                student = await get_or_create_user(
                    session,
                    email=student_email,
                    role=UserRole.STUDENT,
                )
                students[student_email] = student

            # Check if assignment already exists
            result = await session.execute(
                select(TeacherStudent).where(
                    TeacherStudent.teacher_id == teacher.id,
                    TeacherStudent.student_id == student.id,
                )
            )
            existing = result.scalar_one_or_none()

            if not existing:
                # Create direct teacher-student assignment
                assignment = TeacherStudent(
                    teacher_id=teacher.id,
                    student_id=student.id,
                )
                session.add(assignment)
                created += 1

        except Exception as e:
            errors.append(f"Row {i}: {teacher_email} -> {student_email} - {e}")

    wb.close()

    print(f"  Created: {created} individual teacher-student assignments")
    if errors:
        print(f"  Errors: {len(errors)}")
        for err in errors[:5]:
            print(f"    {err}")

    return created


async def main():
    """Main import function."""
    print("=" * 60)
    print("Starting data import...")
    print("=" * 60)

    # File paths (mounted in Docker container)
    students_file = "/data/База ученики.xlsx"
    individuals_file = "/data/Индивиды.xlsx"
    groups_file = "/data/База групп.csv"

    # Check files exist
    for f in [students_file, individuals_file, groups_file]:
        if not Path(f).exists():
            print(f"ERROR: File not found: {f}")
            return

    async with async_session_maker() as session:
        try:
            # 1. Import students
            students = await import_students(session, students_file)

            # 2. Extract teachers from groups CSV
            teachers = await import_teachers_from_groups(session, groups_file)

            # 3. Extract teachers from individuals file
            teachers = await import_teachers_from_individuals(session, individuals_file, teachers)

            # 4. Import groups
            groups = await import_groups(session, groups_file, teachers, students)

            # 5. Create individual teacher-student assignments
            individuals = await create_individual_assignments(session, individuals_file, teachers, students)

            # Commit all changes
            await session.commit()

            print("\n" + "=" * 60)
            print("Import completed successfully!")
            print("=" * 60)
            print(f"  Total students: {len(students)}")
            print(f"  Total teachers: {len(teachers)}")
            print(f"  Total groups: {len(groups)}")
            print(f"  Individual assignments: {individuals}")

        except Exception as e:
            await session.rollback()
            print(f"\nERROR: Import failed: {e}")
            import traceback
            traceback.print_exc()
            raise


if __name__ == "__main__":
    asyncio.run(main())
