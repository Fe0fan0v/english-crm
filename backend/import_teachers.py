"""Import teachers from Excel file to database."""
import asyncio
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import select

# Add backend to path
sys.path.insert(0, str(Path(__file__).parent))

from app.database import async_session_maker
from app.models import Level, User, UserRole
from app.security import hash_password


async def import_teachers(excel_path: str):
    """Import teachers from Excel file."""
    print(f"Reading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Read headers
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")

    # Read all rows (skip header)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"Found {len(rows)} teachers in Excel")

    async with async_session_maker() as db:
        # Get all levels
        levels_result = await db.execute(select(Level))
        levels = {level.name: level for level in levels_result.scalars().all()}
        print(f"Available levels: {list(levels.keys())}")

        # Get existing users
        users_result = await db.execute(select(User))
        existing_emails = {user.email for user in users_result.scalars().all()}
        print(f"Existing users: {len(existing_emails)}")

        created_count = 0
        skipped_count = 0
        updated_count = 0

        for idx, row in enumerate(rows, start=2):
            name = row[0]  # Имя
            email = row[1]  # Email
            password = row[2]  # Пароль
            # row[3] is empty
            level_name = row[4] if len(row) > 4 else None  # Уровень

            if not name or not email:
                print(f"Row {idx}: Skipping - missing name or email")
                skipped_count += 1
                continue

            # Use default password if not provided
            if not password:
                password = "teacher123"
            else:
                password = str(password)

            # Check if user exists
            user_result = await db.execute(
                select(User).where(User.email == email.strip().lower())
            )
            existing_user = user_result.scalar_one_or_none()

            if existing_user:
                print(f"Row {idx}: User {email} already exists, skipping")
                skipped_count += 1
                continue

            # Map Excel level names to database levels
            level_id = None
            if level_name:
                level_str = str(level_name).lower()
                # Map "12+ лет" -> "12+ месяцев", "6-12 лет" -> "6-12 месяцев"
                if "12+" in level_str:
                    # Find "12+ месяцев" level
                    for level_obj in levels.values():
                        if "12+" in level_obj.name.lower():
                            level_id = level_obj.id
                            break
                elif "6-12" in level_str or "6-11" in level_str:
                    # Find "6-12 месяцев" level
                    for level_obj in levels.values():
                        if "6-12" in level_obj.name.lower():
                            level_id = level_obj.id
                            break
                elif "1-6" in level_str or "1-5" in level_str:
                    # Find "1-6 месяцев" level
                    for level_obj in levels.values():
                        if "1-6" in level_obj.name.lower():
                            level_id = level_obj.id
                            break

            # Create new teacher
            hashed_password = hash_password(password)
            new_user = User(
                name=name.strip(),
                email=email.strip().lower(),
                password_hash=hashed_password,
                role=UserRole.TEACHER,
                level_id=level_id,
                balance="0",
                is_active=True,
            )
            db.add(new_user)
            created_count += 1
            print(
                f"Row {idx}: Created teacher {name} ({email}), "
                f"level: {level_id}, password: {password}"
            )

        # Commit all changes
        await db.commit()
        print(f"\n=== Import Summary ===")
        print(f"Created: {created_count}")
        print(f"Skipped (already exists): {skipped_count}")
        print(f"Total processed: {len(rows)}")


if __name__ == "__main__":
    excel_file = r"C:\Users\cypre\Downloads\Telegram Desktop\Преподы.xlsx"
    asyncio.run(import_teachers(excel_file))
